from __future__ import annotations

import argparse
import base64
import csv
import datetime as dt
import html
import json
import os
import re
import sys
import urllib.error
import urllib.request
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from email import policy
from email.parser import BytesParser
from email.utils import parsedate_to_datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable

SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]
APP_DIR = Path(__file__).resolve().parent
DEFAULT_QUERY = (
    'from:(no-reply@info.patreon.com) '
    '("새로 가입했습니다" OR "회원으로 가입했습니다" OR "joined as" OR "new member")'
)

DEFAULT_CONFIG: dict[str, Any] = {
    "gmail_query": DEFAULT_QUERY,
    "tiers_usd": {"1": "2.99", "2": "4.99", "3": "9.99", "4": "29.99"},
    "local_price_map": {
        "USD": {"2.99": 1, "4.99": 2, "9.99": 3, "29.99": 4},
        "CAD": {"4.50": 1, "7.50": 2, "15.00": 3, "45.00": 4},
        "EUR": {"2.70": 1, "4.50": 2, "9.00": 3, "27.00": 4},
    },
    "fallback_rates_to_usd": {
        "AUD": "0.66",
        "BRL": "0.18",
        "CAD": "0.72",
        "CHF": "1.10",
        "CNY": "0.14",
        "EUR": "1.08",
        "GBP": "1.25",
        "HKD": "0.13",
        "JPY": "0.0065",
        "KRW": "0.00073",
        "MXN": "0.055",
        "NZD": "0.60",
        "PHP": "0.017",
        "SGD": "0.76",
        "THB": "0.027",
    },
    "max_relative_tier_error": "0.35",
    "excluded_emails": [],
}

CURRENCY_ALIASES: list[tuple[str, str]] = [
    ("US$", "USD"),
    ("USD", "USD"),
    ("CA$", "CAD"),
    ("CAD", "CAD"),
    ("C$", "CAD"),
    ("AU$", "AUD"),
    ("A$", "AUD"),
    ("AUD", "AUD"),
    ("NZ$", "NZD"),
    ("NZD", "NZD"),
    ("HK$", "HKD"),
    ("HKD", "HKD"),
    ("SGD", "SGD"),
    ("S$", "SGD"),
    ("MX$", "MXN"),
    ("MXN", "MXN"),
    ("BRL", "BRL"),
    ("R$", "BRL"),
    ("CHF", "CHF"),
    ("CNY", "CNY"),
    ("RMB", "CNY"),
    ("PHP", "PHP"),
    ("THB", "THB"),
    ("KRW", "KRW"),
    ("JPY", "JPY"),
    ("€", "EUR"),
    ("EUR", "EUR"),
    ("£", "GBP"),
    ("GBP", "GBP"),
    ("¥", "JPY"),
    ("₩", "KRW"),
    ("$", "USD"),
]

CURRENCY_TOKEN_RE = "|".join(re.escape(token) for token, _ in CURRENCY_ALIASES)
AMOUNT_RE = r"[0-9][0-9,]*(?:[.][0-9]+)?|[0-9]+(?:,[0-9]{1,2})"
MONEY_RE = re.compile(
    rf"(?P<currency>{CURRENCY_TOKEN_RE})\s*(?P<amount>{AMOUNT_RE})",
    re.IGNORECASE,
)
MONEY_FRAGMENT_RE = rf"(?:{CURRENCY_TOKEN_RE})\s*(?:{AMOUNT_RE})"
EMAIL_RE = re.compile(r"[\w.!#$%&'*+/=?^_`{|}~-]+@[\w.-]+\.[A-Za-z]{2,}")

JOIN_TERMS = (
    "새로 가입했습니다",
    "회원으로 가입했습니다",
    "joined as",
    "new member",
    "has joined",
    "joined your membership",
)

CSV_FIELDS = [
    "received_at",
    "member_name",
    "member_email",
    "tier",
    "tier_usd",
    "original_amount",
    "currency",
    "amount",
    "usd_estimate",
    "match_method",
    "confidence",
    "subject",
    "from",
    "gmail_id",
    "message_id",
]


@dataclass
class Money:
    currency: str
    amount: Decimal
    raw: str


@dataclass
class SourceMessage:
    gmail_id: str
    message_id: str
    subject: str
    sender: str
    received_at: dt.datetime | None
    text: str


@dataclass
class MemberRecord:
    received_at: str
    member_name: str
    member_email: str
    money: Money
    tier: int | None
    tier_usd: Decimal | None
    usd_estimate: Decimal | None
    match_method: str
    confidence: str
    subject: str
    sender: str
    gmail_id: str
    message_id: str

    def as_row(self) -> dict[str, str]:
        return {
            "received_at": self.received_at,
            "member_name": self.member_name,
            "member_email": self.member_email,
            "tier": "" if self.tier is None else str(self.tier),
            "tier_usd": decimal_to_str(self.tier_usd),
            "original_amount": self.money.raw,
            "currency": self.money.currency,
            "amount": decimal_to_str(self.money.amount),
            "usd_estimate": decimal_to_str(self.usd_estimate),
            "match_method": self.match_method,
            "confidence": self.confidence,
            "subject": self.subject,
            "from": self.sender,
            "gmail_id": self.gmail_id,
            "message_id": self.message_id,
        }


class HtmlTextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        value = data.strip()
        if value:
            self.parts.append(value)

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() in {"br", "p", "div", "li", "tr", "h1", "h2", "h3"}:
            self.parts.append("\n")

    def get_text(self) -> str:
        return normalize_text(" ".join(self.parts))


class RateProvider:
    def __init__(
        self,
        cache_path: Path,
        fallback_rates: dict[str, Decimal],
        exchange_mode: str,
    ) -> None:
        self.cache_path = cache_path
        self.fallback_rates = fallback_rates
        self.exchange_mode = exchange_mode
        self.cache = self._load_cache()

    def rate_to_usd(self, currency: str, received_at: dt.datetime | None) -> tuple[Decimal | None, str]:
        currency = currency.upper()
        if currency == "USD":
            return Decimal("1"), "usd"

        if self.exchange_mode != "off":
            date_key = "latest"
            if self.exchange_mode == "email-date" and received_at:
                date_key = received_at.date().isoformat()
            cache_key = f"{date_key}:{currency}:USD"
            if cache_key in self.cache:
                return Decimal(str(self.cache[cache_key])), "frankfurter-cache"
            fetched = self._fetch_rate(currency, date_key)
            if fetched is not None:
                self.cache[cache_key] = str(fetched)
                self._save_cache()
                return fetched, "frankfurter"

        if currency in self.fallback_rates:
            return self.fallback_rates[currency], "fallback-config"
        return None, "missing-rate"

    def _fetch_rate(self, currency: str, date_key: str) -> Decimal | None:
        url = f"https://api.frankfurter.app/{date_key}?from={currency}&to=USD"
        try:
            with urllib.request.urlopen(url, timeout=10) as response:
                payload = json.loads(response.read().decode("utf-8"))
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError):
            return None
        try:
            return Decimal(str(payload["rates"]["USD"]))
        except (KeyError, InvalidOperation):
            return None

    def _load_cache(self) -> dict[str, str]:
        if not self.cache_path.exists():
            return {}
        try:
            return json.loads(self.cache_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return {}

    def _save_cache(self) -> None:
        self.cache_path.parent.mkdir(parents=True, exist_ok=True)
        self.cache_path.write_text(
            json.dumps(self.cache, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )


def main(argv: list[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    config = load_config(Path(args.config))
    out_path = Path(args.out)
    rate_provider = RateProvider(
        cache_path=Path(args.cache),
        fallback_rates=parse_decimal_map(config.get("fallback_rates_to_usd", {})),
        exchange_mode=args.exchange_mode,
    )

    if args.eml_dir:
        messages = list(read_eml_dir(Path(args.eml_dir), args.limit))
    else:
        query = build_query(args.query or config["gmail_query"], args.after, args.before)
        messages = read_gmail_messages(
            credentials_path=Path(args.credentials),
            token_path=Path(args.token),
            query=query,
            limit=args.limit,
            include_spam_trash=args.include_spam_trash,
        )

    excluded_emails = {email.lower() for email in config.get("excluded_emails", [])}
    records: list[MemberRecord] = []
    seen = set()
    for message in messages:
        dedupe_key = message.message_id or message.gmail_id or f"{message.subject}:{message.received_at}"
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        record = parse_member_record(message, config, rate_provider, excluded_emails)
        if record is not None:
            records.append(record)

    records.sort(key=lambda item: item.received_at, reverse=True)
    if args.dry_run:
        print_summary(records, out_path=None, xlsx_path=None)
        return 0

    write_csv(out_path, records)
    xlsx_path = Path(args.xlsx) if args.xlsx else None
    if xlsx_path:
        write_xlsx(xlsx_path, records)
    html_path = Path(args.html) if args.html else None
    if html_path:
        write_html_report(html_path, records)
    print_summary(records, out_path=out_path, xlsx_path=xlsx_path, html_path=html_path)
    return 0


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Export new Patreon member emails from Gmail to CSV/XLSX.",
    )
    parser.add_argument("--credentials", default=str(APP_DIR / "credentials.json"))
    parser.add_argument("--token", default=str(APP_DIR / "token.json"))
    parser.add_argument("--config", default=str(APP_DIR / "config.json"))
    parser.add_argument("--cache", default=str(APP_DIR / ".cache" / "rates.json"))
    parser.add_argument("--query", help="Override Gmail search query.")
    parser.add_argument("--after", help="Append Gmail after: filter, e.g. 2026/01/01.")
    parser.add_argument("--before", help="Append Gmail before: filter, e.g. 2026/06/01.")
    parser.add_argument("--limit", type=int, default=0, help="0 means no explicit limit.")
    parser.add_argument("--out", default=str(APP_DIR / "output" / "patreon_members.csv"))
    parser.add_argument("--xlsx", help="Optional XLSX output path.")
    parser.add_argument("--html", help="Optional HTML report output path.")
    parser.add_argument("--eml-dir", help="Parse exported .eml files instead of Gmail API.")
    parser.add_argument(
        "--exchange-mode",
        choices=["email-date", "latest", "off"],
        default="email-date",
        help="Use historical date, latest online rate, or only config fallback rates.",
    )
    parser.add_argument("--include-spam-trash", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    return parser


def load_config(config_path: Path) -> dict[str, Any]:
    if not config_path.exists():
        example_path = config_path.with_name("config.example.json")
        if example_path.exists():
            config_path.write_text(example_path.read_text(encoding="utf-8"), encoding="utf-8")
        else:
            config_path.write_text(
                json.dumps(DEFAULT_CONFIG, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
    raw = json.loads(config_path.read_text(encoding="utf-8"))
    merged = dict(DEFAULT_CONFIG)
    merged.update(raw)
    return merged


def parse_decimal_map(raw: dict[str, str]) -> dict[str, Decimal]:
    parsed: dict[str, Decimal] = {}
    for key, value in raw.items():
        try:
            parsed[key.upper()] = Decimal(str(value))
        except InvalidOperation:
            continue
    return parsed


def build_query(base_query: str, after: str | None, before: str | None) -> str:
    parts = [base_query.strip()]
    if after:
        parts.append(f"after:{normalize_gmail_date(after)}")
    if before:
        parts.append(f"before:{normalize_gmail_date(before)}")
    return " ".join(parts)


def normalize_gmail_date(value: str) -> str:
    return value.strip().replace("-", "/")


def read_gmail_messages(
    credentials_path: Path,
    token_path: Path,
    query: str,
    limit: int,
    include_spam_trash: bool,
) -> Iterable[SourceMessage]:
    service = build_gmail_service(credentials_path, token_path)
    fetched = 0
    page_token = None
    while True:
        page_size = 500
        if limit:
            page_size = max(1, min(500, limit - fetched))
        response = (
            service.users()
            .messages()
            .list(
                userId="me",
                q=query,
                maxResults=page_size,
                pageToken=page_token,
                includeSpamTrash=include_spam_trash,
            )
            .execute()
        )
        for item in response.get("messages", []):
            raw_message = (
                service.users()
                .messages()
                .get(userId="me", id=item["id"], format="full")
                .execute()
            )
            yield source_from_gmail(raw_message)
            fetched += 1
            if limit and fetched >= limit:
                return
        page_token = response.get("nextPageToken")
        if not page_token:
            return


def build_gmail_service(credentials_path: Path, token_path: Path) -> Any:
    if not credentials_path.exists():
        raise SystemExit(
            f"Missing {credentials_path}. Create a Gmail API Desktop OAuth client, "
            "download it, and save it as credentials.json."
        )
    try:
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise SystemExit(
            "Missing Google API libraries. Run: python -m pip install -r requirements.txt"
        ) from exc

    creds = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(credentials_path), SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.write_text(creds.to_json(), encoding="utf-8")
    return build("gmail", "v1", credentials=creds)


def source_from_gmail(message: dict[str, Any]) -> SourceMessage:
    payload = message.get("payload", {})
    headers = headers_to_dict(payload.get("headers", []))
    received_at = None
    if message.get("internalDate"):
        received_at = dt.datetime.fromtimestamp(int(message["internalDate"]) / 1000, tz=dt.timezone.utc)
    elif headers.get("date"):
        received_at = parse_email_date(headers["date"])
    return SourceMessage(
        gmail_id=message.get("id", ""),
        message_id=headers.get("message-id", ""),
        subject=headers.get("subject", ""),
        sender=headers.get("from", ""),
        received_at=received_at,
        text=extract_payload_text(payload),
    )


def headers_to_dict(headers: list[dict[str, str]]) -> dict[str, str]:
    return {item.get("name", "").lower(): item.get("value", "") for item in headers}


def extract_payload_text(payload: dict[str, Any]) -> str:
    plain_parts: list[str] = []
    html_parts: list[str] = []
    for part in walk_payload_parts(payload):
        mime_type = part.get("mimeType", "")
        data = part.get("body", {}).get("data")
        if not data:
            continue
        decoded = decode_base64url(data)
        if mime_type == "text/plain":
            plain_parts.append(decoded)
        elif mime_type == "text/html":
            html_parts.append(html_to_text(decoded))
    text = "\n".join(plain_parts or html_parts)
    return normalize_text(text)


def walk_payload_parts(payload: dict[str, Any]) -> Iterable[dict[str, Any]]:
    if payload.get("parts"):
        for part in payload["parts"]:
            yield from walk_payload_parts(part)
    else:
        yield payload


def decode_base64url(data: str) -> str:
    padded = data + "=" * (-len(data) % 4)
    return base64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8", errors="replace")


def read_eml_dir(eml_dir: Path, limit: int) -> Iterable[SourceMessage]:
    paths = sorted(eml_dir.glob("*.eml"))
    if limit:
        paths = paths[:limit]
    for path in paths:
        yield source_from_eml(path)


def source_from_eml(path: Path) -> SourceMessage:
    with path.open("rb") as handle:
        message = BytesParser(policy=policy.default).parse(handle)
    subject = str(message.get("subject", ""))
    sender = str(message.get("from", ""))
    received_at = parse_email_date(str(message.get("date", "")))
    message_id = str(message.get("message-id", ""))

    plain = ""
    html_value = ""
    if message.is_multipart():
        for part in message.walk():
            content_type = part.get_content_type()
            if content_type == "text/plain" and not plain:
                plain = part.get_content()
            elif content_type == "text/html" and not html_value:
                html_value = part.get_content()
    else:
        if message.get_content_type() == "text/html":
            html_value = message.get_content()
        else:
            plain = message.get_content()
    return SourceMessage(
        gmail_id=path.stem,
        message_id=message_id,
        subject=subject,
        sender=sender,
        received_at=received_at,
        text=normalize_text(plain or html_to_text(html_value)),
    )


def parse_email_date(value: str) -> dt.datetime | None:
    if not value:
        return None
    try:
        parsed = parsedate_to_datetime(value)
    except (TypeError, ValueError):
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=dt.timezone.utc)
    return parsed


def parse_member_record(
    message: SourceMessage,
    config: dict[str, Any],
    rate_provider: RateProvider,
    excluded_emails: set[str],
) -> MemberRecord | None:
    search_text = normalize_text(f"{message.subject}\n{message.text}")
    if not is_join_message(search_text):
        return None
    money = find_money(search_text)
    if money is None:
        return None
    name = extract_member_name(search_text) or ""
    member_email = extract_member_email(search_text, name, excluded_emails)
    tier, tier_usd, usd_estimate, method, confidence = classify_tier(
        money=money,
        received_at=message.received_at,
        config=config,
        rate_provider=rate_provider,
    )
    return MemberRecord(
        received_at=format_datetime(message.received_at),
        member_name=name,
        member_email=member_email,
        money=money,
        tier=tier,
        tier_usd=tier_usd,
        usd_estimate=usd_estimate,
        match_method=method,
        confidence=confidence,
        subject=message.subject,
        sender=message.sender,
        gmail_id=message.gmail_id,
        message_id=message.message_id,
    )


def is_join_message(text: str) -> bool:
    lower = text.lower()
    return any(term.lower() in lower for term in JOIN_TERMS)


def find_money(text: str) -> Money | None:
    match = MONEY_RE.search(text)
    if not match:
        return None
    currency_token = match.group("currency")
    amount_raw = match.group("amount")
    currency = normalize_currency(currency_token)
    amount = parse_amount(amount_raw)
    if currency is None or amount is None:
        return None
    return Money(currency=currency, amount=amount, raw=f"{currency_token}{amount_raw}")


def normalize_currency(token: str) -> str | None:
    normalized = token.upper()
    for alias, currency in CURRENCY_ALIASES:
        if normalized == alias.upper():
            return currency
    return None


def parse_amount(value: str) -> Decimal | None:
    cleaned = value.strip().replace(" ", "")
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        head, tail = cleaned.rsplit(",", 1)
        if len(tail) in {1, 2}:
            cleaned = f"{head}.{tail}"
        else:
            cleaned = cleaned.replace(",", "")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def extract_member_name(text: str) -> str | None:
    patterns = [
        rf"새로운\s+{MONEY_FRAGMENT_RE}\s+회원!\s*(?:[^\w가-힣ぁ-んァ-ン一-龥]+\s*)?(?P<name>.+?)님이\s+새로\s+가입했습니다",
        rf"(?P<name>[^\n!]+?)님이\s+{MONEY_FRAGMENT_RE}\s+회원(?:으로)?\s+가입했습니다",
        r"(?:^|\n)\s*(?P<name>[^\n]+?)님이\s+새로\s+가입했습니다",
        rf"New\s+{MONEY_FRAGMENT_RE}\s+member!\s*(?P<name>.+?)\s+(?:joined|has joined)",
        rf"(?P<name>.+?)\s+(?:joined|has joined)\s+as\s+a\s+{MONEY_FRAGMENT_RE}\s+member",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return clean_name(match.group("name"))
    return None


def clean_name(value: str) -> str:
    value = re.sub(r"^[^\w가-힣ぁ-んァ-ン一-龥]+", "", value.strip())
    value = re.sub(r"\s+", " ", value)
    value = value.strip(" -:|")
    return value[:120]


def extract_member_email(text: str, name: str, excluded_emails: set[str]) -> str:
    candidates: list[str] = []
    if name:
        idx = text.find(name)
        if idx >= 0:
            candidates.extend(EMAIL_RE.findall(text[idx : idx + 700]))
    candidates.extend(EMAIL_RE.findall(text))
    for candidate in candidates:
        normalized = candidate.lower()
        if normalized not in excluded_emails and "patreon.com" not in normalized:
            return candidate
    return ""


def classify_tier(
    money: Money,
    received_at: dt.datetime | None,
    config: dict[str, Any],
    rate_provider: RateProvider,
) -> tuple[int | None, Decimal | None, Decimal | None, str, str]:
    tiers = {int(k): Decimal(str(v)) for k, v in config["tiers_usd"].items()}
    local_match = match_local_price(money, config.get("local_price_map", {}))
    if local_match is not None:
        tier_usd = tiers.get(local_match)
        usd_estimate = tier_usd if money.currency != "USD" else money.amount
        return local_match, tier_usd, usd_estimate, "local_price_map", "high"

    rate, rate_method = rate_provider.rate_to_usd(money.currency, received_at)
    if rate is None:
        return None, None, None, rate_method, "needs_review"
    usd_estimate = money.amount * rate
    tier, tier_usd, rel_error = nearest_tier(usd_estimate, tiers)
    max_error = Decimal(str(config.get("max_relative_tier_error", "0.35")))
    if rel_error <= max_error:
        confidence = "medium" if rel_error <= Decimal("0.20") else "low"
        return tier, tier_usd, usd_estimate, rate_method, confidence
    return None, None, usd_estimate, rate_method, "needs_review"


def match_local_price(money: Money, local_price_map: dict[str, dict[str, int]]) -> int | None:
    candidates = local_price_map.get(money.currency.upper(), {})
    for price_raw, tier in candidates.items():
        price = Decimal(str(price_raw))
        tolerance = Decimal("0.03") if price < Decimal("100") else Decimal("1")
        if abs(money.amount - price) <= tolerance:
            return int(tier)
    return None


def nearest_tier(usd_estimate: Decimal, tiers: dict[int, Decimal]) -> tuple[int, Decimal, Decimal]:
    best_tier = min(tiers, key=lambda tier: abs(usd_estimate - tiers[tier]))
    tier_usd = tiers[best_tier]
    rel_error = abs(usd_estimate - tier_usd) / tier_usd
    return best_tier, tier_usd, rel_error


def html_to_text(value: str) -> str:
    extractor = HtmlTextExtractor()
    extractor.feed(html.unescape(value))
    return extractor.get_text()


def normalize_text(value: str) -> str:
    value = html.unescape(value)
    value = value.replace("\xa0", " ")
    value = re.sub(r"[ \t\r\f\v]+", " ", value)
    value = re.sub(r"\n\s+", "\n", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def format_datetime(value: dt.datetime | None) -> str:
    if value is None:
        return ""
    return value.astimezone().isoformat(timespec="seconds")


def decimal_to_str(value: Decimal | None) -> str:
    if value is None:
        return ""
    quant = Decimal("0.01") if abs(value) >= Decimal("0.01") else Decimal("0.0001")
    return str(value.quantize(quant))


def write_csv(path: Path, records: list[MemberRecord]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for record in records:
            writer.writerow(record.as_row())


def write_xlsx(path: Path, records: list[MemberRecord]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit("Missing openpyxl. Run: python -m pip install -r requirements.txt") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Patreon members"
    sheet.append(CSV_FIELDS)
    for record in records:
        row = record.as_row()
        sheet.append([row[field] for field in CSV_FIELDS])
    for col_index, field in enumerate(CSV_FIELDS, start=1):
        width = max(len(field), *(len(str(row.as_row()[field])) for row in records)) if records else len(field)
        sheet.column_dimensions[get_column_letter(col_index)].width = min(max(width + 2, 10), 50)
    workbook.save(path)


def write_html_report(path: Path, records: list[MemberRecord]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = [record.as_row() for record in records]
    tier_counts = {"1": 0, "2": 0, "3": 0, "4": 0, "needs_review": 0}
    for row in rows:
        key = row["tier"] if row["tier"] in {"1", "2", "3", "4"} else "needs_review"
        tier_counts[key] += 1
    max_count = max(tier_counts.values(), default=1) or 1
    generated_at = dt.datetime.now().astimezone().isoformat(timespec="seconds")
    total = len(rows)
    needs_review = tier_counts["needs_review"]
    cards = "\n".join(
        f"""
        <section class="metric">
          <span>{escape_html(label)}</span>
          <strong>{value}</strong>
        </section>
        """
        for label, value in [
            ("Total", total),
            ("Tier 1", tier_counts["1"]),
            ("Tier 2", tier_counts["2"]),
            ("Tier 3", tier_counts["3"]),
            ("Tier 4", tier_counts["4"]),
            ("Needs review", needs_review),
        ]
    )
    bars = "\n".join(
        f"""
        <div class="bar-row">
          <span class="bar-label">{escape_html(label)}</span>
          <div class="bar-track"><div class="bar-fill" style="width: {count / max_count * 100:.1f}%"></div></div>
          <span class="bar-count">{count}</span>
        </div>
        """
        for label, count in [
            ("Tier 1", tier_counts["1"]),
            ("Tier 2", tier_counts["2"]),
            ("Tier 3", tier_counts["3"]),
            ("Tier 4", tier_counts["4"]),
            ("Needs review", needs_review),
        ]
    )
    table_rows = "\n".join(
        f"""
        <tr>
          <td>{escape_html(row["received_at"])}</td>
          <td>{escape_html(row["member_name"])}</td>
          <td>{escape_html(row["member_email"])}</td>
          <td>{escape_html(row["tier"] or "review")}</td>
          <td>{escape_html(row["original_amount"])}</td>
          <td>{escape_html(row["usd_estimate"])}</td>
          <td>{escape_html(row["confidence"])}</td>
        </tr>
        """
        for row in rows
    )
    document = f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Patreon Member Report</title>
  <style>
    :root {{
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --ink: #172033;
      --muted: #667085;
      --line: #d9dee8;
      --accent: #0f766e;
      --accent-soft: #ccfbf1;
      --warn: #a16207;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--ink);
      font-family: "Segoe UI", Arial, sans-serif;
      line-height: 1.45;
    }}
    main {{
      max-width: 1180px;
      margin: 0 auto;
      padding: 28px;
    }}
    header {{
      display: flex;
      justify-content: space-between;
      gap: 18px;
      align-items: end;
      margin-bottom: 22px;
    }}
    h1 {{
      margin: 0 0 6px;
      font-size: 28px;
      letter-spacing: 0;
    }}
    .muted {{ color: var(--muted); }}
    .metrics {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
      gap: 12px;
      margin-bottom: 16px;
    }}
    .metric, .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
    }}
    .metric {{
      padding: 16px;
      min-height: 92px;
    }}
    .metric span {{
      display: block;
      color: var(--muted);
      font-size: 13px;
      margin-bottom: 10px;
    }}
    .metric strong {{
      font-size: 30px;
      font-weight: 700;
    }}
    .panel {{
      padding: 18px;
      margin-top: 16px;
    }}
    h2 {{
      margin: 0 0 14px;
      font-size: 18px;
      letter-spacing: 0;
    }}
    .bar-row {{
      display: grid;
      grid-template-columns: 110px 1fr 54px;
      align-items: center;
      gap: 12px;
      min-height: 34px;
      margin: 8px 0;
    }}
    .bar-label, .bar-count {{ font-size: 14px; }}
    .bar-count {{ text-align: right; color: var(--muted); }}
    .bar-track {{
      height: 16px;
      background: #eef2f6;
      border-radius: 999px;
      overflow: hidden;
    }}
    .bar-fill {{
      height: 100%;
      background: var(--accent);
      min-width: 2px;
    }}
    .table-wrap {{
      overflow-x: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 920px;
      background: var(--panel);
    }}
    th, td {{
      padding: 10px 12px;
      border-bottom: 1px solid var(--line);
      text-align: left;
      font-size: 13px;
      vertical-align: top;
    }}
    th {{
      background: #f0f3f7;
      color: #344054;
      font-weight: 700;
      position: sticky;
      top: 0;
    }}
    tr:last-child td {{ border-bottom: 0; }}
    @media (max-width: 760px) {{
      main {{ padding: 18px; }}
      header {{ display: block; }}
      .bar-row {{ grid-template-columns: 84px 1fr 42px; }}
    }}
  </style>
</head>
<body>
  <main>
    <header>
      <div>
        <h1>Patreon Member Report</h1>
        <div class="muted">Generated {escape_html(generated_at)}</div>
      </div>
      <div class="muted">CSV and Excel files are in the same output folder.</div>
    </header>

    <div class="metrics">
      {cards}
    </div>

    <section class="panel">
      <h2>Tier Chart</h2>
      {bars}
    </section>

    <section class="panel">
      <h2>Member Table</h2>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>Received</th>
              <th>Name</th>
              <th>Email</th>
              <th>Tier</th>
              <th>Original amount</th>
              <th>USD estimate</th>
              <th>Confidence</th>
            </tr>
          </thead>
          <tbody>
            {table_rows}
          </tbody>
        </table>
      </div>
    </section>
  </main>
</body>
</html>
"""
    path.write_text(document, encoding="utf-8")


def escape_html(value: object) -> str:
    return html.escape(str(value), quote=True)


def print_summary(
    records: list[MemberRecord],
    out_path: Path | None,
    xlsx_path: Path | None,
    html_path: Path | None = None,
) -> None:
    tier_counts: dict[str, int] = {}
    for record in records:
        key = f"tier {record.tier}" if record.tier else "needs_review"
        tier_counts[key] = tier_counts.get(key, 0) + 1
    print(f"Parsed {len(records)} Patreon join records.")
    if tier_counts:
        print("Tier counts: " + ", ".join(f"{key}={value}" for key, value in sorted(tier_counts.items())))
    if out_path:
        print(f"CSV: {out_path}")
    if xlsx_path:
        print(f"XLSX: {xlsx_path}")
    if html_path:
        print(f"HTML: {html_path}")


if __name__ == "__main__":
    raise SystemExit(main())
